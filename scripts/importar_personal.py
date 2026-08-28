"""
Carga el listado de personal desde el Excel de recursos humanos a la hoja
`personal` del Sheet de la plataforma.

Se usa para autocompletar el registro: el inspector escribe su cédula y salen
su nombre, correo, cargo y área. Con eso el nombre queda escrito siempre igual,
que es el problema que arrastran los Sheets de inspección.

NO se importa el correo personal: para lo que hace la plataforma basta el
corporativo, y traer 145 correos personales es exponer datos que no se
necesitan.

La hoja se reemplaza entera en cada corrida, porque la fuente de verdad es el
Excel de RR. HH. Si alguien se retira, se quita de allá y se vuelve a correr.

Uso:
    python scripts/importar_personal.py <service-account.json> <archivo.xlsx>
"""

import io
import re
import sys

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build

SHEET_ID = "1z0Wcc2-A0S6JCVjm_e3kfcXZgZMgr3g_VpOTMq6X3IU"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
HOJA = "personal"

ENCABEZADOS = [
    "cedula",
    "nombre",
    "correo",
    "cargo",
    "categoria",
    "area",
    "area_trabajo",
    "lugar",
]

# Posiciones en el Excel de RR. HH. (0-indexadas).
COL = {
    "cedula": 2,
    "nombre": 3,
    "correo": 5,  # CORREO CORPORATIVO. El personal, columna 4, no se importa.
    "cargo": 6,
    "categoria": 7,
    "area": 8,
    "area_trabajo": 9,
    "lugar": 10,
}


def _texto(fila, i: int) -> str:
    if i >= len(fila) or fila[i] is None:
        return ""
    # Los saltos de línea dentro de una celda rompen la lectura después.
    return re.sub(r"\s+", " ", str(fila[i])).strip()


def main() -> int:
    if len(sys.argv) < 3:
        print(__doc__)
        return 1

    libro = openpyxl.load_workbook(sys.argv[2], data_only=True)
    ws = libro[libro.sheetnames[0]]

    personas = []
    vistas = set()
    for fila in ws.iter_rows(min_row=2, values_only=True):
        cedula = re.sub(r"\D", "", _texto(fila, COL["cedula"]))
        if not cedula or cedula in vistas:
            continue
        vistas.add(cedula)
        personas.append(
            [
                cedula,
                _texto(fila, COL["nombre"]),
                _texto(fila, COL["correo"]).lower(),
                _texto(fila, COL["cargo"]),
                _texto(fila, COL["categoria"]),
                _texto(fila, COL["area"]),
                _texto(fila, COL["area_trabajo"]),
                _texto(fila, COL["lugar"]),
            ]
        )

    print(f"personas leídas del Excel: {len(personas)}")
    sin_correo = sum(1 for p in personas if "@" not in p[2])
    if sin_correo:
        print(f"  aviso: {sin_correo} sin correo corporativo")

    cred = service_account.Credentials.from_service_account_file(sys.argv[1], scopes=SCOPES)
    svc = build("sheets", "v4", credentials=cred, cache_discovery=False)

    meta = svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
    existentes = {h["properties"]["title"]: h["properties"]["sheetId"] for h in meta["sheets"]}

    if HOJA not in existentes:
        svc.spreadsheets().batchUpdate(
            spreadsheetId=SHEET_ID,
            body={"requests": [{"addSheet": {"properties": {"title": HOJA}}}]},
        ).execute()
        print(f'hoja "{HOJA}" creada')
        meta = svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
        existentes = {h["properties"]["title"]: h["properties"]["sheetId"] for h in meta["sheets"]}

    # Se limpia y se vuelve a escribir: es un espejo del Excel, no un histórico.
    svc.spreadsheets().values().clear(
        spreadsheetId=SHEET_ID, range=HOJA, body={}
    ).execute()
    svc.spreadsheets().values().update(
        spreadsheetId=SHEET_ID,
        range=f"{HOJA}!A1",
        valueInputOption="RAW",
        body={"values": [ENCABEZADOS] + personas},
    ).execute()

    svc.spreadsheets().batchUpdate(
        spreadsheetId=SHEET_ID,
        body={
            "requests": [
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": existentes[HOJA],
                            "startRowIndex": 0,
                            "endRowIndex": 1,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "textFormat": {"bold": True},
                                "backgroundColor": {"red": 0.99, "green": 0.95, "blue": 0.95},
                            }
                        },
                        "fields": "userEnteredFormat(textFormat,backgroundColor)",
                    }
                },
                {
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": existentes[HOJA],
                            "gridProperties": {"frozenRowCount": 1},
                        },
                        "fields": "gridProperties.frozenRowCount",
                    }
                },
            ]
        },
    ).execute()

    print(f'hoja "{HOJA}": {len(personas)} personas escritas')
    print("No se importó el correo personal, a propósito.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
